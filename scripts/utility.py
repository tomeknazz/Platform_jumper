import pygame
import os
import sys
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles.builtins import currency

BASE_IMG_PATH = "data/images/"


# Załadowanie obrazka
def load_image(path):
    img = pygame.image.load(BASE_IMG_PATH + path).convert_alpha()
    return img


# Załadowanie wielu obrazków
def load_images(path):
    images = []
    for img_name in sorted(os.listdir(BASE_IMG_PATH + path)):
        images.append(load_image(path + "/" + img_name))
    return images

# Zapisywanie do xlsx
def save_to_excel(user_name, time, total_jumps, level_beaten):
    filename = 'Ranking.xlsx'

    try:
        # Try to load the existing workbook
        wb = load_workbook(filename)
        ws = wb.active
    except FileNotFoundError:
        # If the workbook does not exist, create a new one
        wb = Workbook()
        ws = wb.active
        ws.title = "Ranking"
        ws.append(['User Name', 'Time', 'Total Jumps', "Map"])

    # Append the data
    ws.append([user_name, time, total_jumps, level_beaten])

    wb.save(filename)

from openpyxl import load_workbook

from openpyxl import load_workbook

def load_from_excel(current_map):
    filename = 'Ranking.xlsx'
    try:
        wb = load_workbook(filename)
        ws = wb.active
        data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row.count(None) == len(row):
                continue
            if len(row) > 3 and row[3] == current_map:
                # Pomijaj 4. kolumnę w danych (indeks 3)
                filtered_row = row[:3] + row[4:]  # zachowaj wszystko oprócz indeksu 3
                data.append(filtered_row)

        # Sortuj według drugiej kolumny (czas – indeks 1)
        data.sort(key=lambda x: x[1])
        return data

    except FileNotFoundError:
        print("File not found. Please save some data first.")
        return []
    except Exception as e:
        print(f"An error occurred: {e}")
        return []


    except FileNotFoundError:
        print("File not found. Please save some data first.")
        return []
    except Exception as e:
        print(f"An error occurred: {e}")
        return []

# Klasa z animacjami
class Animation:
    def __init__(self, images, img_dur=5, loop=True):
        # Atrybuty
        self.images = images
        self.loop = loop
        self.img_duration = img_dur
        self.done = False
        self.frame = 0

    def copy(self):
        return Animation(self.images, self.img_duration, self.loop)

    # Update animacji
    def update(self):
        if self.loop:
            self.frame = (self.frame + 1) % (self.img_duration * len(self.images))
        else:
            self.frame = min(self.frame + 1, self.img_duration * len(self.images) - 1)
            if self.frame >= self.img_duration * len(self.images) - 1:
                self.done = True

    def img(self):
        return self.images[int(self.frame / self.img_duration)]
